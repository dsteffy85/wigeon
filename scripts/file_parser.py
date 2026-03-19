#!/usr/bin/env python3
"""
WIGEON File Parser
Handles multiple file formats: Excel (.xls, .xlsx), CSV, XML, ZIP
"""

import csv
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

# Excel parsing
try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not available - Excel parsing will be limited")

try:
    import xlrd

    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    print("⚠️  xlrd not available - .xls parsing not supported")


class FileParser:
    """Multi-format file parser for WIGEON"""

    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self.file_format = self.file_path.suffix.lower()

        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

    def parse(self):
        """Parse file based on format"""
        if self.file_format in [".xlsx", ".xlsm"]:
            return self.parse_xlsx()
        elif self.file_format == ".xls":
            return self.parse_xls()
        elif self.file_format == ".csv":
            return self.parse_csv()
        elif self.file_format == ".xml":
            return self.parse_xml()
        elif self.file_format == ".zip":
            return self.parse_zip()
        else:
            raise ValueError(f"Unsupported file format: {self.file_format}")

    def parse_csv(self):
        """Parse CSV file"""
        rows = []
        headers = []

        with open(self.file_path, encoding="utf-8-sig") as f:
            # Try to detect delimiter
            sample = f.read(4096)
            f.seek(0)

            # Check for common delimiters
            if sample.count("|") > sample.count(","):
                delimiter = "|"
            elif sample.count("\t") > sample.count(","):
                delimiter = "\t"
            else:
                delimiter = ","

            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []

            for row in reader:
                # Clean up None values
                cleaned_row = {k: v for k, v in row.items() if k is not None}
                rows.append(cleaned_row)

        return {
            "format": "csv",
            "sheets": {"Sheet1": {"headers": headers, "rows": rows, "row_count": len(rows)}},
            "sheet_count": 1,
            "total_rows": len(rows),
        }

    def parse_xlsx(self):
        """Parse Excel .xlsx file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for .xlsx files. Install: pip install openpyxl")

        workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        results = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value is not None else f"Column_{cell.column}")

            # Get data rows
            rows = []
            for _row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        # Convert datetime objects to strings
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        row_data[headers[idx]] = value
                rows.append(row_data)

            results[sheet_name] = {"headers": headers, "rows": rows, "row_count": len(rows)}

        return {
            "format": "xlsx",
            "sheets": results,
            "sheet_count": len(results),
            "total_rows": sum(s["row_count"] for s in results.values()),
        }

    def parse_xls(self):
        """Parse Excel .xls file"""
        if not XLRD_AVAILABLE:
            raise ImportError("xlrd is required for .xls files. Install: pip install xlrd")

        workbook = xlrd.open_workbook(self.file_path)
        results = {}

        for sheet in workbook.sheets():
            # Get headers from first row
            headers = []
            if sheet.nrows > 0:
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(0, col_idx)
                    headers.append(str(cell_value) if cell_value else f"Column_{col_idx + 1}")

            # Get data rows
            rows = []
            for row_idx in range(1, sheet.nrows):
                row_data = {}
                for col_idx in range(sheet.ncols):
                    if col_idx < len(headers):
                        cell_value = sheet.cell_value(row_idx, col_idx)

                        # Handle date cells
                        cell_type = sheet.cell_type(row_idx, col_idx)
                        if cell_type == xlrd.XL_CELL_DATE:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                            cell_value = datetime(*date_tuple).isoformat()

                        row_data[headers[col_idx]] = cell_value
                rows.append(row_data)

            results[sheet.name] = {"headers": headers, "rows": rows, "row_count": len(rows)}

        return {
            "format": "xls",
            "sheets": results,
            "sheet_count": len(results),
            "total_rows": sum(s["row_count"] for s in results.values()),
        }

    def parse_xml(self):
        """Parse XML file"""
        tree = ET.parse(self.file_path)
        root = tree.getroot()

        # Convert XML to list of dictionaries
        rows = []

        def xml_to_dict(element):
            """Recursively convert XML element to dictionary"""
            result = {}

            # Add attributes
            if element.attrib:
                result.update(element.attrib)

            # Add text content
            if element.text and element.text.strip():
                result["_text"] = element.text.strip()

            # Add child elements
            for child in element:
                child_data = xml_to_dict(child)

                # Handle multiple children with same tag
                if child.tag in result:
                    if not isinstance(result[child.tag], list):
                        result[child.tag] = [result[child.tag]]
                    result[child.tag].append(child_data)
                else:
                    result[child.tag] = child_data

            return result

        # Check if root has repeating child elements (common in data exports)
        child_tags = [child.tag for child in root]
        if child_tags and len(set(child_tags)) == 1:
            # All children have same tag - treat as rows
            for child in root:
                rows.append(xml_to_dict(child))
        else:
            # Single record or mixed structure
            rows.append(xml_to_dict(root))

        return {"format": "xml", "root_tag": root.tag, "rows": rows, "row_count": len(rows)}

    def parse_zip(self):
        """Parse ZIP file - extract and parse contents"""
        results = {}
        extracted_files = []

        # Create temp extraction directory
        extract_dir = self.file_path.parent / f"_extracted_{self.file_path.stem}"
        extract_dir.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(self.file_path, "r") as zip_ref:
                zip_ref.extractall(extract_dir)
                extracted_files = list(extract_dir.rglob("*"))

            # Parse each extracted file
            for file_path in extracted_files:
                if file_path.is_file():
                    file_format = file_path.suffix.lower()

                    if file_format in [".xlsx", ".xlsm", ".xls", ".xml", ".csv"]:
                        try:
                            parser = FileParser(file_path)
                            results[file_path.name] = parser.parse()
                        except Exception as e:
                            results[file_path.name] = {"error": str(e), "format": file_format}
                    else:
                        results[file_path.name] = {
                            "format": file_format,
                            "skipped": True,
                            "reason": "Unsupported format",
                        }

        finally:
            # Cleanup extracted files
            import shutil

            if extract_dir.exists():
                shutil.rmtree(extract_dir)

        return {"format": "zip", "files": results, "file_count": len(results)}

    def get_summary(self):
        """Get file summary without parsing full content"""
        return {
            "file_name": self.file_path.name,
            "file_path": str(self.file_path),
            "file_format": self.file_format,
            "file_size": self.file_path.stat().st_size,
            "modified_date": datetime.fromtimestamp(self.file_path.stat().st_mtime).isoformat(),
        }


def normalize_data(parsed_data, third_party_name):
    """
    Normalize parsed data into standard format for database storage
    Returns list of row dictionaries
    """
    normalized_rows = []

    if parsed_data["format"] in ["xlsx", "xls", "csv"]:
        # Excel/CSV files - combine all sheets
        for sheet_name, sheet_data in parsed_data["sheets"].items():
            for row in sheet_data["rows"]:
                normalized_row = {"_sheet": sheet_name, "_third_party": third_party_name, **row}
                normalized_rows.append(normalized_row)

    elif parsed_data["format"] == "xml":
        # XML files
        for row in parsed_data["rows"]:
            normalized_row = {"_third_party": third_party_name, **row}
            normalized_rows.append(normalized_row)

    elif parsed_data["format"] == "zip":
        # ZIP files - process each contained file
        for file_name, file_data in parsed_data["files"].items():
            if "error" in file_data or file_data.get("skipped"):
                continue

            # Recursively normalize nested data
            nested_rows = normalize_data(file_data, third_party_name)
            for row in nested_rows:
                row["_source_file"] = file_name
                normalized_rows.append(row)

    return normalized_rows


if __name__ == "__main__":
    # Test parser
    print("🦆 WIGEON File Parser - Test Mode")
    print("\nSupported formats:")
    print("  ✓ Excel (.xlsx, .xlsm)" if OPENPYXL_AVAILABLE else "  ✗ Excel (.xlsx, .xlsm) - install openpyxl")
    print("  ✓ Excel (.xls)" if XLRD_AVAILABLE else "  ✗ Excel (.xls) - install xlrd")
    print("  ✓ XML (.xml)")
    print("  ✓ ZIP (.zip)")
