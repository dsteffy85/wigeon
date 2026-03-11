#!/usr/bin/env python3
"""
Create sample data files for testing WIGEON
"""

import openpyxl
from datetime import datetime, timedelta
import random
import xml.etree.ElementTree as ET
from pathlib import Path

# Create samples directory
samples_dir = Path(__file__).parent
samples_dir.mkdir(exist_ok=True)

print("🦆 Creating sample data files for WIGEON testing...\n")

# 1. Create sample Excel file - Sales Report
print("1. Creating sample Excel file (sales_report.xlsx)...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sales Data"

# Headers
headers = ["Order ID", "Customer", "Product", "Quantity", "Unit Price", "Total", "Order Date", "Status"]
ws.append(headers)

# Sample data
customers = ["Acme Corp", "TechStart Inc", "Global Solutions", "Innovate LLC", "Future Systems"]
products = ["Widget A", "Widget B", "Gadget X", "Tool Pro", "Device Plus"]
statuses = ["Completed", "Pending", "Shipped", "Processing"]

for i in range(1, 51):
    order_date = datetime.now() - timedelta(days=random.randint(1, 30))
    quantity = random.randint(1, 100)
    unit_price = round(random.uniform(10, 500), 2)
    total = round(quantity * unit_price, 2)
    
    ws.append([
        f"ORD-{1000+i}",
        random.choice(customers),
        random.choice(products),
        quantity,
        unit_price,
        total,
        order_date.strftime("%Y-%m-%d"),
        random.choice(statuses)
    ])

wb.save(samples_dir / "sales_report.xlsx")
print("   ✅ Created sales_report.xlsx (50 rows)")

# 2. Create sample Excel file - Inventory Report
print("\n2. Creating sample Excel file (inventory_report.xlsx)...")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Inventory"

headers2 = ["SKU", "Product Name", "Category", "Quantity on Hand", "Reorder Level", "Unit Cost", "Location"]
ws2.append(headers2)

categories = ["Electronics", "Hardware", "Software", "Accessories", "Tools"]
locations = ["Warehouse A", "Warehouse B", "Distribution Center", "Store 1", "Store 2"]

for i in range(1, 31):
    ws2.append([
        f"SKU-{2000+i}",
        f"Product {chr(65+i%26)}{i}",
        random.choice(categories),
        random.randint(0, 500),
        random.randint(10, 50),
        round(random.uniform(5, 200), 2),
        random.choice(locations)
    ])

wb2.save(samples_dir / "inventory_report.xlsx")
print("   ✅ Created inventory_report.xlsx (30 rows)")

# 3. Create sample XML file
print("\n3. Creating sample XML file (shipments.xml)...")
root = ET.Element("Shipments")
root.set("generated", datetime.now().isoformat())

for i in range(1, 21):
    shipment = ET.SubElement(root, "Shipment")
    
    ET.SubElement(shipment, "ShipmentID").text = f"SHIP-{3000+i}"
    ET.SubElement(shipment, "OrderID").text = f"ORD-{1000+random.randint(1,50)}"
    ET.SubElement(shipment, "Carrier").text = random.choice(["UPS", "FedEx", "DHL", "USPS"])
    ET.SubElement(shipment, "TrackingNumber").text = f"1Z{random.randint(100000000, 999999999)}"
    ET.SubElement(shipment, "Status").text = random.choice(["In Transit", "Delivered", "Out for Delivery", "Pending"])
    
    ship_date = datetime.now() - timedelta(days=random.randint(1, 14))
    ET.SubElement(shipment, "ShipDate").text = ship_date.strftime("%Y-%m-%d")
    
    if random.random() > 0.3:
        delivery_date = ship_date + timedelta(days=random.randint(2, 7))
        ET.SubElement(shipment, "DeliveryDate").text = delivery_date.strftime("%Y-%m-%d")

tree = ET.ElementTree(root)
ET.indent(tree, space="  ")
tree.write(samples_dir / "shipments.xml", encoding="utf-8", xml_declaration=True)
print("   ✅ Created shipments.xml (20 shipments)")

# 4. Create a multi-sheet Excel file
print("\n4. Creating multi-sheet Excel file (quarterly_report.xlsx)...")
wb3 = openpyxl.Workbook()

# Sheet 1: Summary
ws_summary = wb3.active
ws_summary.title = "Summary"
ws_summary.append(["Metric", "Q1", "Q2", "Q3", "Q4"])
ws_summary.append(["Revenue", 125000, 145000, 167000, 189000])
ws_summary.append(["Orders", 450, 520, 610, 680])
ws_summary.append(["Customers", 89, 102, 118, 134])

# Sheet 2: Details
ws_details = wb3.create_sheet("Monthly Details")
ws_details.append(["Month", "Revenue", "Orders", "New Customers", "Repeat Customers"])
months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for month in months:
    ws_details.append([
        month,
        random.randint(30000, 60000),
        random.randint(150, 250),
        random.randint(10, 30),
        random.randint(40, 80)
    ])

wb3.save(samples_dir / "quarterly_report.xlsx")
print("   ✅ Created quarterly_report.xlsx (2 sheets)")

print("\n✅ All sample files created successfully!")
print(f"\nSample files location: {samples_dir}")
print("\nYou can now test WIGEON with these files:")
print("  python wigeon.py ingest --file samples/sales_report.xlsx --third-party 'Acme Corp' --email vendor@acme.com")
